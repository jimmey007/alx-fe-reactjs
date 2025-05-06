from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

# Initialize Selenium WebDriver
driver = webdriver.Chrome()  # Ensure ChromeDriver is in PATH
driver.get("https://example.com/form")  # Replace with your target website

# Load Excel file
workbook = load_workbook("data.xlsx")
sheet = workbook.active

# Read data from Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
    name, email = row[0], row[1]  # Assuming columns: Name, Email

    # Find web form fields and fill them
    driver.find_element(By.ID, "name").send_keys(name)
    driver.find_element(By.ID, "email").send_keys(email)
    driver.find_element(By.ID, "submit").click()

    time.sleep(1)  # Wait for page to reload/reset

# Close the browser
driver.quit()